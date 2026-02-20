#!/usr/bin/env python3
"""
Generate Excel file from user story JSON data.

Usage:
    python generate_excel.py --output stories.xlsx --input stories.json
    python generate_excel.py --output stories.xlsx --data '[{"category_epic": "...", ...}]'

Dependencies:
    pip install openpyxl
"""

import argparse
import json
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def create_user_stories_excel(stories: list, output_path: str) -> str:
    """Create a formatted Excel file from user story data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "User Stories"

    headers = ["Category/Epic", "Title", "User Story", "Acceptance Criteria", "Requirement", "Pain Point", "Created Date"]

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    keys = ["category_epic", "title", "user_story", "acceptance_criteria", "requirement", "pain_point", "created_date"]
    for row_idx, story in enumerate(stories, 2):
        for col, key in enumerate(keys, 1):
            cell = ws.cell(row=row_idx, column=col, value=story.get(key, ""))
            cell.alignment = cell_alignment
            cell.border = thin_border

    # Column widths
    column_widths = {1: 30, 2: 35, 3: 60, 4: 50, 5: 50, 6: 45, 7: 15}
    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description="Generate formatted Excel file from user story JSON data",
        epilog="Example: python generate_excel.py --input stories.json --output user_stories.xlsx"
    )
    parser.add_argument("--output", "-o", required=True, help="Output Excel file path (.xlsx)")
    parser.add_argument("--data", "-d", help="JSON string of user stories")
    parser.add_argument("--input", "-i", help="Path to JSON file containing user stories")

    args = parser.parse_args()

    # Load stories from input source
    if args.input:
        try:
            with open(args.input, "r", encoding="utf-8") as f:
                stories = json.load(f)
        except FileNotFoundError:
            print(f"Error: Input file not found: {args.input}")
            sys.exit(1)
        except json.JSONDecodeError as e:
            print(f"Error: Invalid JSON in input file: {e}")
            sys.exit(1)
    elif args.data:
        try:
            stories = json.loads(args.data)
        except json.JSONDecodeError as e:
            print(f"Error: Invalid JSON data: {e}")
            sys.exit(1)
    else:
        # Read from stdin
        try:
            stories = json.load(sys.stdin)
        except json.JSONDecodeError as e:
            print(f"Error: Invalid JSON from stdin: {e}")
            sys.exit(1)

    if not isinstance(stories, list):
        print("Error: Data must be a JSON array of user story objects")
        sys.exit(1)

    if not stories:
        print("Warning: No stories provided, creating empty file")

    output_path = create_user_stories_excel(stories, args.output)
    print(f"Created: {output_path} ({len(stories)} stories)")


if __name__ == "__main__":
    main()
